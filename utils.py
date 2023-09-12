import re

from openpyxl import load_workbook

def get_data():
    sheet = load_workbook("move-all-accounts.xlsx")["Report"]
    _, _, *data = list(sheet.iter_rows(min_row=2, values_only=True))
    return data


def clean_reason(label_text):
    label_text = re.sub(r"\|НАРЕДИТЕЛ: [А-ЯA-Z0-9 ]*", "", label_text)
    label_text = re.sub(r"\|ПОЛУЧАТЕЛ: [А-ЯA-Z0-9 ]*", "", label_text)
    label_text = re.sub(r"\|СМЕТКА: [A-Z0-9]*", "", label_text)
    label_text = re.sub(r"\|BIC: [A-Z]*", "", label_text)
    label_text = re.sub(r" КУРС: [0-9.]*", "", label_text)
    label_text = re.sub(r"  BG Karta \*\*\*\d{4}", "", label_text)
    label_text = re.sub(r" [0-9|]+\*\*\*[0-9|]+", "", label_text)
    label_text = re.sub(r" [0-9|]+\.[0-9|]+\.[0-9|]+ [0-9|]+:[0-9|]+:[0-9|]+", "", label_text)
    label_text = re.sub(r"\s{2,}", " ", label_text)
    label_text = label_text.replace("|", "") # this should hapen as a last replacement
    label_text = ";".join(list(set(label_text.split(";"))))
    return label_text